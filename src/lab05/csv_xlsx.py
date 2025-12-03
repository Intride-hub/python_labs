import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    csv_p = Path(csv_path)
    xlsx_p = Path(xlsx_path)

    if not csv_p.exists():
        raise FileNotFoundError(f"Файл не найден: {csv_path}")

    # --- чтение CSV ---
    with csv_p.open(encoding="utf-8") as f:
        reader = csv.reader(f)
        data = list(reader)

    if len(data) == 0:
        raise ValueError("CSV пустой")

    # --- создание XLSX ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # запись строк
    for row in data:
        ws.append(row)

    # авто-ширина колонок
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        adj_width = max(max_len, 8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = adj_width

    wb.save(xlsx_p)
